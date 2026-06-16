"""
split.py  –  EPC Automation v6 | Reliance Retail EPC Dept
----------------------------------------------------------
Reads the Master Excel file and generates one .xlsx per PM
with formatting matching the original master file exactly.

HOW TO RUN:  Double-click 1_RUN_SPLIT.bat  OR  python split.py
"""

import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
with open("config.json") as f:
    cfg = json.load(f)

MASTER_FILE   = cfg["master_file"]
MASTER_SHEET  = cfg["master_sheet"]
UNIQUE_KEY    = cfg["unique_key"]
OUTPUT_FOLDER = cfg["output_folder"]
PM_LIST       = cfg["pms"]

print()
print("=" * 60)
print("  EPC AUTOMATION  -  SPLIT  (v6)")
print("  Reliance Retail | EPC Department")
print("=" * 60)
print()

if not os.path.exists(MASTER_FILE):
    print(f"\n  ERROR: '{MASTER_FILE}' not found.")
    input("\n  Press Enter to close...")
    raise SystemExit

print(f"  Reading: {MASTER_FILE}  (sheet: '{MASTER_SHEET}') ...")
df = pd.read_excel(MASTER_FILE, sheet_name=MASTER_SHEET, dtype=str)
df.columns  = df.columns.str.strip()
df["PM"]    = df["PM"].str.strip()
df[UNIQUE_KEY] = df[UNIQUE_KEY].str.strip()
df = df.fillna("")
print(f"  Loaded {len(df):,} rows  |  {len(df.columns)} columns\n")

os.makedirs(OUTPUT_FOLDER, exist_ok=True)
datestamp = datetime.today().strftime("%d-%m-%Y")

# ── Formatting (matching original) ───────────────────────────────────────────
HDR_FONT     = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FONT_DARK = Font(name="Calibri", size=10, bold=True, color="000000")  # for light bg
BODY_FONT    = Font(name="Calibri", size=10)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=False)
T            = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=T, right=T, top=T, bottom=T)

FILL_DEFAULT = PatternFill(FILL_SOLID, fgColor="538135")   # green
FILL_THEME4  = PatternFill(FILL_SOLID, fgColor="BDD7EE")   # light blue
FILL_THEME1  = PatternFill(FILL_SOLID, fgColor="000000")   # black
FILL_THEME0  = PatternFill(FILL_SOLID, fgColor="538135")   # green (same as default)
FILL_ODD     = PatternFill(FILL_SOLID, fgColor="FFFFFF")
FILL_EVEN    = PatternFill(FILL_SOLID, fgColor="F2F2F2")

THEME4_COLS  = {"PM Head","PM","PM Planner","AOP / NON AOP","Planned (Month) Bucket","Actual (Month) Bucket"}
THEME1_COLS  = {"EPC Status (Current Week)","% Completion (Current Week)","Remarks (Current Week)"}
THEME0_COLS  = {"Target Date","Target","P vs A","Dehire Target","Need Push / Sure","Remarks",
                "RFP","Rent Start Date","Dead Rent Days","Dead Rent Bucket","Dead Rent Status",
                "ML Date","LOA Date","EPC Status (Last Week)","% Completion (Last Week)",
                "Remarks (Last Week)","AOP / NON Phase","Phase Wise Sites"}

def hdr_fill(col_name):
    if col_name in THEME4_COLS: return FILL_THEME4
    if col_name in THEME1_COLS: return FILL_THEME1
    if col_name in THEME0_COLS: return FILL_THEME0
    return FILL_DEFAULT

# ── Split loop ────────────────────────────────────────────────────────────────
print(f"  {'PM':<28}  {'Sites':>6}  File")
print(f"  {'-'*28}  {'-'*6}  {'-'*38}")

results = []
for pm in PM_LIST:
    pm_df = df[df["PM"] == pm].reset_index(drop=True)

    if len(pm_df) == 0:
        print(f"  !  {pm:<26}  {'':>6}  No rows found — check spelling in config.json")
        results.append((pm, 0, None))
        continue

    safe    = pm.replace(" ", "_").replace("/", "-")
    outpath = os.path.join(OUTPUT_FOLDER, f"{safe}_{datestamp}.xlsx")
    cols    = list(pm_df.columns)
    n_cols  = len(cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "My Sites"
    ws.sheet_properties.tabColor = "2E4057"

    # Header row
    LIGHT_BG_COLS = {"PM Head","PM","PM Planner","AOP / NON AOP",
                     "Planned (Month) Bucket","Actual (Month) Bucket"}
    for ci, col_name in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col_name)
        c.font      = HDR_FONT_DARK if col_name in LIGHT_BG_COLS else HDR_FONT
        c.fill      = hdr_fill(col_name)
        c.alignment = CENTER
        c.border    = BORDER
    ws.row_dimensions[1].height = 60

    # Data rows
    for ri, (_, row_data) in enumerate(pm_df.iterrows(), 2):
        fill = FILL_EVEN if ri % 2 == 0 else FILL_ODD
        for ci, col_name in enumerate(cols, 1):
            val = row_data[col_name]
            c = ws.cell(row=ri, column=ci,
                        value="" if str(val) in ("nan","NaT","None") else val)
            c.font      = BODY_FONT
            c.alignment = LEFT
            c.border    = BORDER
            c.fill      = fill

    # Column widths matching original (20.73 default, 40.73 for Site Name)
    for ci, col_name in enumerate(cols, 1):
        cl = get_column_letter(ci)
        ws.column_dimensions[cl].width = 40.73 if col_name == "Site Name" else 20.73

    # Freeze top row + auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    # READ ME sheet
    rws = wb.create_sheet("READ ME")
    rws.sheet_properties.tabColor = "C00000"
    for row_data in [
        ["EPC AUTOMATION - INSTRUCTIONS"],[""],
        [f"PM Name      : {pm}"],
        [f"Total Sites  : {len(pm_df)}"],
        [f"Generated on : {datetime.today().strftime('%d %b %Y')}"],[""],
        ["INSTRUCTIONS:"],
        ["  1. Go to the My Sites tab."],
        ["  2. Update the columns as required."],
        ["  3. Do NOT change the APEX ID column."],
        ["  4. Save the file (Ctrl+S) when done."],
        ["  5. Return the file to the PM_Files folder on shared drive."],
    ]:
        rws.append(row_data)
    rws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    rws["A1"].fill = PatternFill(FILL_SOLID, fgColor="2E4057")
    rws.column_dimensions["A"].width = 60
    rws.row_dimensions[1].height = 30
    rws.sheet_view.showGridLines = False

    wb.active = wb["My Sites"]
    wb.save(outpath)

    print(f"  OK {pm:<26}  {len(pm_df):>6}  {outpath}")
    results.append((pm, len(pm_df), outpath))

ok    = [r for r in results if r[1] > 0]
total = sum(r[1] for r in ok)
print()
print("-" * 60)
print(f"  {len(ok)} PM files created  |  {total:,} sites distributed")
print(f"  Folder: {OUTPUT_FOLDER}\\")
print("-" * 60)
print()
input("  Press Enter to close...")
