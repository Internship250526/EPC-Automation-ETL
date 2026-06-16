"""
merge.py  -  EPC Automation v7 | Reliance Retail EPC Dept
----------------------------------------------------------
Fixes in v7:
  1. Deletions detected and written back
  2. Hidden characters / non-breaking spaces cleaned
  3. Date formats normalised (12-06-2026 == 12/06/2026 == June 12 2026)
  4. Case-insensitive comparison
  5. APEX ID tampered → warning logged
  6. Unknown APEX ID → visible warning, not silent skip
  7. Extra rows in PM file → flagged
  8. Date cells written as clean DD-MM-YYYY strings (no 00:00:00)

HOW TO RUN:  Double-click 2_RUN_MERGE.bat  OR  python merge.py
"""

import pandas as pd
import json
import os
import re
import unicodedata
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.utils import get_column_letter
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
with open("config.json") as f:
    cfg = json.load(f)

MASTER_FILE   = cfg["master_file"]
MASTER_SHEET  = cfg["master_sheet"]
UNIQUE_KEY    = cfg["unique_key"]
OUTPUT_FOLDER = cfg["output_folder"]

print()
print("=" * 60)
print("  EPC AUTOMATION  -  MERGE  (v7)")
print("  Reliance Retail | EPC Department")
print("=" * 60)
print()

# ── Master file check ─────────────────────────────────────────────────────────
if not os.path.exists(MASTER_FILE):
    print(f"  ERROR: '{MASTER_FILE}' not found.")
    input("\n  Press Enter to close...")
    raise SystemExit

stamp = datetime.today().strftime("%d-%m-%Y_%H%M%S")

# ── Helper: clean a cell value ────────────────────────────────────────────────
def clean(val):
    """Strip whitespace, normalise unicode, remove hidden chars."""
    if val is None:
        return ""
    s = str(val).strip()
    # Normalise unicode (e.g. non-breaking space \xa0 → regular space)
    s = unicodedata.normalize("NFKC", s)
    # Remove zero-width characters
    s = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', s)
    s = s.strip()
    if s.lower() in ("nan", "nat", "none", ""):
        return ""
    return s

# ── Helper: normalise date strings ───────────────────────────────────────────
DATE_PATTERNS = [
    "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y",
    "%Y-%m-%d", "%Y/%m/%d",
    "%d-%b-%Y", "%d %b %Y", "%d %B %Y",
    "%d-%m-%y", "%d/%m/%y",
]

def normalise_date(val):
    """Try to parse val as a date and return DD-MM-YYYY string, or return val as-is."""
    # Strip time portion if present (e.g. "2026-06-12 00:00:00")
    v = re.sub(r'\s+\d{2}:\d{2}:\d{2}$', '', val).strip()
    for fmt in DATE_PATTERNS:
        try:
            dt = datetime.strptime(v, fmt)
            return dt.strftime("%d-%m-%Y")
        except ValueError:
            continue
    return val  # not a date — return as-is

def normalise(val):
    """Clean + normalise dates + lowercase for comparison."""
    v = clean(val)
    if not v:
        return ""
    v = normalise_date(v)
    return v.lower()   # case-insensitive comparison

def display_val(val):
    """Clean value for writing to output (preserve case, normalise date)."""
    v = clean(val)
    if not v:
        return ""
    return normalise_date(v)

# ── Load master ───────────────────────────────────────────────────────────────
print(f"  Reading master: {MASTER_FILE} ...")
master_df = pd.read_excel(MASTER_FILE, sheet_name=MASTER_SHEET, dtype=str)
master_df.columns = master_df.columns.str.strip()
master_df[UNIQUE_KEY] = master_df[UNIQUE_KEY].str.strip()
master_df = master_df.fillna("")

# Store master APEX IDs for fast lookup
master_apex_set = set(master_df[UNIQUE_KEY].values)
master_apex_index = {v: i for i, v in enumerate(master_df[UNIQUE_KEY].values)}

print(f"  Loaded {len(master_df):,} rows\n")

# ── Collect PM files ──────────────────────────────────────────────────────────
if not os.path.exists(OUTPUT_FOLDER):
    print(f"  ERROR: '{OUTPUT_FOLDER}' folder not found. Run split.py first.")
    input("\n  Press Enter to close...")
    raise SystemExit

pm_files = sorted([f for f in os.listdir(OUTPUT_FOLDER)
                   if f.endswith(".xlsx") and not f.startswith("~$")])

if not pm_files:
    print(f"  ERROR: No .xlsx files found in '{OUTPUT_FOLDER}'.")
    input("\n  Press Enter to close...")
    raise SystemExit

print(f"  Found {len(pm_files)} PM file(s):\n")

# ── Merge logic ───────────────────────────────────────────────────────────────
pending       = {}   # apex_id → {col: {value, source, conflict}}
conflict_log  = []
warning_log   = []   # unknown IDs, extra rows, tampered IDs

for filename in pm_files:
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    print(f"  Processing: {filename}")

    try:
        wb_check   = openpyxl.load_workbook(filepath, read_only=True)
        sheet_name = "My Sites" if "My Sites" in wb_check.sheetnames else wb_check.sheetnames[0]
        wb_check.close()
        pm_df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
    except Exception as e:
        print(f"    WARNING: Could not read — {e}\n")
        continue

    pm_df.columns = pm_df.columns.str.strip()

    # ── Check APEX ID column exists ───────────────────────────────────────────
    if UNIQUE_KEY not in pm_df.columns:
        print(f"    WARNING: '{UNIQUE_KEY}' column missing — skipping\n")
        warning_log.append({"File": filename, "Issue": f"'{UNIQUE_KEY}' column missing"})
        continue

    pm_df[UNIQUE_KEY] = pm_df[UNIQUE_KEY].str.strip()
    pm_df = pm_df.fillna("")
    cols_present = [c for c in pm_df.columns if c != UNIQUE_KEY]

    updated  = 0
    warnings = 0

    # Check for renamed columns
    master_cols  = set(master_df.columns)
    pm_cols      = set(pm_df.columns) - {UNIQUE_KEY}
    renamed_cols = pm_cols - master_cols
    if renamed_cols:
        for rc in renamed_cols:
            msg = f"Column '{rc}' not found in master — may have been renamed or added"
            warning_log.append({"File": filename, "Issue": msg})
            warnings += 1
        print()
        print("  !" * 30)
        print(f"  CRITICAL WARNING in file: {filename}")
        for rc in renamed_cols:
            print(f"    Column '{rc}' does not exist in master.")
            print(f"    It may have been RENAMED or ADDED by the PM.")
            print(f"    Any data entered under this column will NOT be merged.")
        print("  !" * 30)
        print()
        confirm = input("  Type CONTINUE to proceed anyway, or close this window to abort: ")
        if confirm != "CONTINUE":
            print()
            print("  Merge aborted by admin. Fix the PM file and run again.")
            print()
            input("  Press Enter to close...")
            raise SystemExit

    # Check for duplicate APEX IDs in this PM file
    seen_in_file = {}
    dup_ids = []
    for row_num, (_, row) in enumerate(pm_df.iterrows(), start=2):
        apex_id = clean(row[UNIQUE_KEY])
        if apex_id and apex_id in seen_in_file:
            msg = f"APEX ID '{apex_id}' appears more than once (rows {seen_in_file[apex_id]} and {row_num}) — only first occurrence used"
            warning_log.append({"File": filename, UNIQUE_KEY: apex_id, "Issue": msg})
            dup_ids.append((apex_id, seen_in_file[apex_id], row_num))
            warnings += 1
        elif apex_id:
            seen_in_file[apex_id] = row_num

    if dup_ids:
        print()
        print("  !" * 30)
        print(f"  CRITICAL WARNING in file: {filename}")
        for apex_id, r1, r2 in dup_ids:
            print(f"    APEX ID '{apex_id}' is duplicated (rows {r1} and {r2}).")
            print(f"    Only row {r1} will be used. Row {r2} will be ignored.")
        print("  !" * 30)
        print()
        confirm = input("  Type CONTINUE to proceed anyway, or close this window to abort: ")
        if confirm != "CONTINUE":
            print()
            print("  Merge aborted by admin. Fix the PM file and run again.")
            print()
            input("  Press Enter to close...")
            raise SystemExit

    for row_num, (_, row) in enumerate(pm_df.iterrows(), start=2):
        apex_id = clean(row[UNIQUE_KEY])

        # ── Skip truly empty rows ─────────────────────────────────────────────
        if not apex_id:
            continue

        # ── Skip duplicate rows (already processed first occurrence) ─────────
        first_occurrence = seen_in_file.get(apex_id)
        if first_occurrence and first_occurrence != row_num and apex_id in pending:
            continue

        # ── Detect extra/unknown APEX IDs ────────────────────────────────────
        if apex_id not in master_apex_set:
            msg = f"APEX ID '{apex_id}' (row {row_num}) not found in master"
            print(f"    WARNING: {msg}")
            warning_log.append({"File": filename, UNIQUE_KEY: apex_id,
                                 "Row": row_num, "Issue": msg})
            warnings += 1
            continue

        master_idx = master_apex_index[apex_id]

        if apex_id not in pending:
            pending[apex_id] = {}

        for col in cols_present:
            new_raw = clean(row.get(col, ""))
            new_display = display_val(new_raw)   # clean + date normalised
            new_norm    = normalise(new_raw)      # for comparison (lowercase)

            old_raw     = clean(master_df.at[master_idx, col])
            old_norm    = normalise(old_raw)

            # Skip if both empty
            if new_norm == "" and old_norm == "":
                continue

            # Skip if effectively same value (case-insensitive, date-normalised)
            if new_norm == old_norm:
                continue

            # Genuine change (edit or deletion)
            if col in pending[apex_id]:
                prev = pending[apex_id][col]
                if prev["norm"] != new_norm:
                    # Two different PMs changed the same cell → conflict
                    conflict_log.append({
                        "APEX ID"        : apex_id,
                        "Column"         : col,
                        "First value"    : prev["value"],
                        "First source"   : prev["source"],
                        "Conflict value" : new_display,
                        "Conflict source": filename,
                        "Resolution"     : "First value kept"
                    })
                    pending[apex_id][col]["conflict"] = True
            else:
                pending[apex_id][col] = {
                    "value"   : new_display,
                    "norm"    : new_norm,
                    "source"  : filename,
                    "conflict": False
                }
            updated += 1

    status_msg = f"    OK  {updated} change(s) queued"
    if warnings:
        status_msg += f"  |  {warnings} warning(s) — see log"
    print(status_msg + "\n")

# ── Apply to master_df ────────────────────────────────────────────────────────
cells_changed  = 0
cells_conflict = 0
changed_cells_map = {}

for apex_id, col_changes in pending.items():
    idx = master_apex_index.get(apex_id)
    if idx is None:
        continue
    for col, info in col_changes.items():
        if col not in master_df.columns:
            continue
        master_df.at[idx, col] = info["value"]
        changed_cells_map[(apex_id, col)] = "conflict" if info["conflict"] else "updated"
        if info["conflict"]: cells_conflict += 1
        cells_changed += 1

print(f"  Applying changes ...")

# ── Formatting constants ──────────────────────────────────────────────────────
HDR_FONT      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FONT_DARK = Font(name="Calibri", size=10, bold=True, color="000000")
BODY_FONT     = Font(name="Calibri", size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=False)
T             = Side(style="thin", color="BFBFBF")
BORDER        = Border(left=T, right=T, top=T, bottom=T)

FILL_DEFAULT  = PatternFill(FILL_SOLID, fgColor="538135")
FILL_THEME4   = PatternFill(FILL_SOLID, fgColor="BDD7EE")
FILL_THEME1   = PatternFill(FILL_SOLID, fgColor="000000")
FILL_THEME0   = PatternFill(FILL_SOLID, fgColor="538135")
FILL_ODD      = PatternFill(FILL_SOLID, fgColor="FFFFFF")
FILL_EVEN     = PatternFill(FILL_SOLID, fgColor="F2F2F2")
FILL_UPDATED  = PatternFill(FILL_SOLID, fgColor="FFFF00")
FILL_CONFLICT = PatternFill(FILL_SOLID, fgColor="FFB3B3")

THEME4_COLS   = {"PM Head","PM","PM Planner","AOP / NON AOP",
                 "Planned (Month) Bucket","Actual (Month) Bucket"}
THEME1_COLS   = {"EPC Status (Current Week)","% Completion (Current Week)",
                 "Remarks (Current Week)"}
THEME0_COLS   = {"Target Date","Target","P vs A","Dehire Target","Need Push / Sure",
                 "Remarks","RFP","Rent Start Date","Dead Rent Days","Dead Rent Bucket",
                 "Dead Rent Status","ML Date","LOA Date","EPC Status (Last Week)",
                 "% Completion (Last Week)","Remarks (Last Week)",
                 "AOP / NON Phase","Phase Wise Sites"}
LIGHT_BG_COLS = {"PM Head","PM","PM Planner","AOP / NON AOP",
                 "Planned (Month) Bucket","Actual (Month) Bucket"}

def hdr_fill(col_name):
    if col_name in THEME4_COLS: return FILL_THEME4
    if col_name in THEME1_COLS: return FILL_THEME1
    if col_name in THEME0_COLS: return FILL_THEME0
    return FILL_DEFAULT

# ── Write output workbook ─────────────────────────────────────────────────────
cols   = list(master_df.columns)
n_cols = len(cols)

updated_apex_ids  = {aid for (aid, col), st in changed_cells_map.items() if st == "updated"}
conflict_apex_ids = {aid for (aid, col), st in changed_cells_map.items() if st == "conflict"}

wb = Workbook()
ws = wb.active
ws.title = MASTER_SHEET
ws.sheet_properties.tabColor = "2E4057"

# Header row
for ci, col_name in enumerate(cols, 1):
    c = ws.cell(row=1, column=ci, value=col_name)
    c.font      = HDR_FONT_DARK if col_name in LIGHT_BG_COLS else HDR_FONT
    c.fill      = hdr_fill(col_name)
    c.alignment = CENTER
    c.border    = BORDER
ws.row_dimensions[1].height = 60

# Data rows
for ri, (_, row_data) in enumerate(master_df.iterrows(), 2):
    apex_id   = clean(row_data[UNIQUE_KEY])
    base_fill = FILL_EVEN if ri % 2 == 0 else FILL_ODD

    apex_fill = (FILL_CONFLICT if apex_id in conflict_apex_ids
                 else FILL_UPDATED if apex_id in updated_apex_ids
                 else base_fill)

    for ci, col_name in enumerate(cols, 1):
        raw = row_data[col_name]
        val = display_val(str(raw))   # clean + strip 00:00:00 from dates
        c = ws.cell(row=ri, column=ci, value=val)
        c.font      = BODY_FONT
        c.alignment = LEFT
        c.border    = BORDER
        c.fill      = apex_fill if col_name == UNIQUE_KEY else base_fill

# Column widths + freeze + filter
for ci, col_name in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(ci)].width = (
        40.73 if col_name == "Site Name" else 20.73)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

output_file = f"Updated_{MASTER_FILE}"
wb.save(output_file)

print(f"  Saved  ->  {output_file}")
print(f"  Cells updated  : {cells_changed}")
print(f"  Conflicts      : {cells_conflict}")

# ── Warning / conflict log ────────────────────────────────────────────────────
all_logs = conflict_log + warning_log
if all_logs:
    log_path = f"Merge_Log_{stamp}.xlsx"
    pd.DataFrame(all_logs).to_excel(log_path, index=False)
    print(f"\n  Log saved  ->  {log_path}")
    if conflict_log:
        print(f"  {len(conflict_log)} conflict(s) — RED in updated file")
    if warning_log:
        print(f"  {len(warning_log)} warning(s) — unknown/extra rows in PM files")
else:
    print("\n  No conflicts or warnings. Clean merge!")

print()
print("-" * 60)
print("  MERGE COMPLETE")
print(f"  Open '{output_file}' to see updated master.")
print("  YELLOW = updated row  |  RED = conflict row")
print("-" * 60)
print()
input("  Press Enter to close...")
